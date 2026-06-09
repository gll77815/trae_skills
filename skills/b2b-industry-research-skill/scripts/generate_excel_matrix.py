#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成Excel版竞品-场景交叉矩阵

功能：
1. 生成单表汇总所有场景数据
2. 应用条件格式（市场占有率颜色、斑马纹）
3. 冻结窗格便于浏览
4. 添加筛选和排序功能

使用时机：Phase 2 - 竞品-场景交叉分析
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from datetime import datetime


def create_excel_matrix(industry, scenes, competitors, output_path=None):
    """
    创建Excel版竞品-场景交叉矩阵
    
    参数:
        industry: 行业名称
        scenes: 场景列表，每项包含name, type, pain_points, market_share
        competitors: 竞品列表，每项包含name, category
        output_path: 输出路径（可选）
    
    返回:
        生成的文件路径
    """
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "竞品场景矩阵"
    
    # 定义颜色填充
    fill_high = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # 红色-极高
    fill_medium = PatternFill(start_color="FFB347", end_color="FFB347", fill_type="solid")  # 橙色-中等
    fill_low = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")  # 蓝色-低
    fill_header = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # 深蓝-表头
    fill_zebra = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # 斑马纹
    
    # 定义字体
    font_header = Font(bold=True, color="FFFFFF", size=11)
    font_normal = Font(size=10)
    font_bold = Font(bold=True, size=10)
    
    # 定义边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 定义对齐
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # 写入标题行
    headers = ["场景名称", "场景类型", "场景特点/痛点", "面向客户", "资质要求", "场景可行性"] + \
              [c["name"] for c in competitors]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_center
        cell.border = thin_border
    
    # 写入数据行
    for row_idx, scene in enumerate(scenes, 2):
        # 基础信息列
        ws.cell(row=row_idx, column=1, value=scene.get("name", "")).alignment = align_left
        ws.cell(row=row_idx, column=2, value=scene.get("type", "")).alignment = align_center
        ws.cell(row=row_idx, column=3, value=scene.get("pain_points", "")).alignment = align_left
        ws.cell(row=row_idx, column=4, value=scene.get("target_customers", "")).alignment = align_left
        ws.cell(row=row_idx, column=5, value=scene.get("qualification", "")).alignment = align_center
        ws.cell(row=row_idx, column=6, value=scene.get("feasibility", "")).alignment = align_center
        
        # 竞品市场占有率列
        market_shares = scene.get("market_shares", {})
        for col_idx, competitor in enumerate(competitors, 7):
            share = market_shares.get(competitor["name"], "")
            cell = ws.cell(row=row_idx, column=col_idx, value=share)
            cell.alignment = align_center
            
            # 根据市场占有率设置颜色
            if share:
                share_lower = str(share).lower()
                if any(x in share_lower for x in ["极高", "70-90%", "70%", "80%", "90%"]):
                    cell.fill = fill_high
                elif any(x in share_lower for x in ["中", "中等", "30-50%", "40%", "50%"]):
                    cell.fill = fill_medium
                elif any(x in share_lower for x in ["低", "低等", "10-20%", "10%", "20%", "少"]):
                    cell.fill = fill_low
        
        # 应用斑马纹
        if row_idx % 2 == 0:
            for col in range(1, len(headers) + 1):
                if not ws.cell(row=row_idx, column=col).fill or \
                   ws.cell(row=row_idx, column=col).fill.fill_type != "solid":
                    ws.cell(row=row_idx, column=col).fill = fill_zebra
        
        # 应用边框
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).border = thin_border
    
    # 设置列宽
    column_widths = {
        1: 25,  # 场景名称
        2: 12,  # 场景类型
        3: 40,  # 场景特点/痛点
        4: 20,  # 面向客户
        5: 12,  # 资质要求
        6: 12,  # 场景可行性
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # 竞品列宽度
    for col_idx in range(7, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15
    
    # 设置行高
    ws.row_dimensions[1].height = 30
    for row_idx in range(2, len(scenes) + 2):
        ws.row_dimensions[row_idx].height = 40
    
    # 冻结首行
    ws.freeze_panes = "A2"
    
    # 添加筛选
    ws.auto_filter.ref = ws.dimensions
    
    # 添加说明Sheet
    ws_help = wb.create_sheet("使用说明")
    help_content = [
        ["竞品-场景交叉矩阵使用说明", ""],
        ["", ""],
        ["颜色说明", ""],
        ["红色", "市场占有率极高（70-90%）- 红海市场，谨慎进入"],
        ["橙色", "市场占有率中等（30-50%）- 竞争激烈"],
        ["蓝色", "市场占有率低（10-20%）- 潜在机会"],
        ["空白", "无明确产品或市场空白"],
        ["", ""],
        ["场景可行性标注", ""],
        ["可切入", "有市场空间且团队有能力切入"],
        ["PASS", "竞品太强或市场太小，不建议进入"],
        ["观望", "政策不确定或技术不成熟，暂缓"],
        ["", ""],
        ["操作提示", ""],
        ["1", "点击表头可筛选特定类型的场景"],
        ["2", "在'场景可行性'列标注您的判断"],
        ["3", "使用Excel排序功能按场景类型分组查看"],
        ["", ""],
        ["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
    ]
    
    for row_idx, row_data in enumerate(help_content, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_help.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
            elif row_idx in [3, 9, 15]:
                cell.font = Font(bold=True, size=11)
    
    ws_help.column_dimensions["A"].width = 15
    ws_help.column_dimensions["B"].width = 50
    
    # 保存文件
    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d")
        output_path = f"{industry}行业_竞品场景矩阵_汇总表_{timestamp}.xlsx"
    
    wb.save(output_path)
    print(f"Excel矩阵已生成: {output_path}")
    return output_path


def main():
    """示例用法"""
    # 示例数据
    scenes = [
        {
            "name": "智能巡检",
            "type": "生产运营",
            "pain_points": "人工巡检效率低、漏检率高",
            "target_customers": "大型制造企业",
            "qualification": "无",
            "feasibility": "可切入",
            "market_shares": {"厂商A": "极高(70-80%)", "厂商B": "中(30-40%)"}
        },
        {
            "name": "设备预测性维护",
            "type": "生产运营",
            "pain_points": "设备故障导致停产损失大",
            "target_customers": "设备密集型企业",
            "qualification": "有",
            "feasibility": "观望",
            "market_shares": {"厂商A": "低(10-20%)", "厂商C": "中(40-50%)"}
        },
    ]
    
    competitors = [
        {"name": "厂商A", "category": "行业龙头系"},
        {"name": "厂商B", "category": "互联网系"},
        {"name": "厂商C", "category": "第三方专业厂商"},
    ]
    
    create_excel_matrix("示例行业", scenes, competitors, "示例_竞品场景矩阵.xlsx")


if __name__ == "__main__":
    main()
