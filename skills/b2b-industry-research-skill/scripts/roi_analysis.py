#!/usr/bin/env python3
"""
ROI分析工具
生成客户视角和己方视角的投入产出分析Excel
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def create_customer_roi_excel(scenarios, output_path):
    """
    生成客户视角ROI分析Excel
    
    Args:
        scenarios: 场景列表，每个场景为字典
            {name, customer_investment, customer_annual_benefit, payback_period, rating}
        output_path: 输出文件路径
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "客户视角ROI分析"
    
    # 样式定义
    HDR_FILL = PatternFill("solid", fgColor="1F4E79")
    HDR_FONT = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
    GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
    GREEN_FONT = Font(name="Microsoft YaHei", size=9, bold=True, color="006100")
    YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
    YELLOW_FONT = Font(name="Microsoft YaHei", size=9, bold=True, color="9C6500")
    ORANGE_FILL = PatternFill("solid", fgColor="FCE4D6")
    ORANGE_FONT = Font(name="Microsoft YaHei", size=9, bold=True, color="BF4000")
    BORDER = Border(
        left=Side("thin", "D9DEE7"), right=Side("thin", "D9DEE7"),
        top=Side("thin", "D9DEE7"), bottom=Side("thin", "D9DEE7")
    )
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # 标题
    ws.merge_cells("A1:F1")
    ws["A1"] = "客户视角投入产出分析"
    ws["A1"].font = Font(name="Microsoft YaHei", size=14, bold=True, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    
    # 表头
    headers = ["场景名称", "客户投入(万元)", "客户年收益(万元)", "投资回收期", "5年ROI", "评级"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = BORDER
    
    # 数据行
    for idx, scenario in enumerate(scenarios, 3):
        row_data = [
            scenario['name'],
            scenario.get('customer_investment', 0),
            scenario.get('customer_annual_benefit', 0),
            scenario.get('payback_period', ''),
            scenario.get('roi_5year', ''),
            scenario.get('rating', '')
        ]
        
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=idx, column=col_idx, value=val)
            cell.border = BORDER
            cell.alignment = CENTER
            
            if col_idx in [2, 3] and isinstance(val, (int, float)):
                cell.number_format = '#,##0'
        
        # 评级颜色
        rating = scenario.get('rating', '')
        rating_cell = ws.cell(row=idx, column=6)
        if "★★★★★" in rating:
            rating_cell.fill = GREEN_FILL
            rating_cell.font = GREEN_FONT
        elif "★★★★" in rating:
            rating_cell.fill = YELLOW_FILL
            rating_cell.font = YELLOW_FONT
        elif "★★★" in rating:
            rating_cell.fill = ORANGE_FILL
            rating_cell.font = ORANGE_FONT
    
    # 设置列宽
    ws.column_dimensions["A"].width = 20
    for col in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 15
    
    wb.save(output_path)
    print(f"客户视角ROI分析已保存：{output_path}")


def create_vendor_roi_excel(scenarios, output_path):
    """
    生成己方视角ROI分析Excel
    
    Args:
        scenarios: 场景列表，每个场景为字典
            {name, rd_team_size, rd_duration, rd_investment, 
             delivery_cost, quote, gross_margin, reusable_ratio,
             target_customers_3y, revenue_3y, profit_3y, payback_period, rating}
        output_path: 输出文件路径
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "己方视角ROI分析"
    
    # 样式
    HDR_FILL = PatternFill("solid", fgColor="1F4E79")
    HDR_FONT = Font(name="Microsoft YaHei", size=9, bold=True, color="FFFFFF")
    GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
    GREEN_FONT = Font(name="Microsoft YaHei", size=9, bold=True, color="006100")
    YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
    YELLOW_FONT = Font(name="Microsoft YaHei", size=9, bold=True, color="9C6500")
    BORDER = Border(
        left=Side("thin", "D9DEE7"), right=Side("thin", "D9DEE7"),
        top=Side("thin", "D9DEE7"), bottom=Side("thin", "D9DEE7")
    )
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # 标题
    ws.merge_cells("A1:N1")
    ws["A1"] = "己方（产品研发方）视角投入产出分析"
    ws["A1"].font = Font(name="Microsoft YaHei", size=14, bold=True, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    
    # 表头
    headers = [
        "场景名称", "研发团队", "研发周期(月)", "研发总投入",
        "单项目交付成本", "单项目报价", "毛利率", "可复用占比",
        "3年目标客户", "3年总营收", "3年总利润", "研发投入回收期", "评级"
    ]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = BORDER
    
    # 数据行
    for idx, scenario in enumerate(scenarios, 3):
        row_data = [
            scenario['name'],
            scenario.get('rd_team_size', 0),
            scenario.get('rd_duration', 0),
            scenario.get('rd_investment', 0),
            scenario.get('delivery_cost', 0),
            scenario.get('quote', 0),
            scenario.get('gross_margin', ''),
            scenario.get('reusable_ratio', ''),
            scenario.get('target_customers_3y', 0),
            scenario.get('revenue_3y', 0),
            scenario.get('profit_3y', 0),
            scenario.get('payback_period', ''),
            scenario.get('rating', '')
        ]
        
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=idx, column=col_idx, value=val)
            cell.border = BORDER
            cell.alignment = CENTER
            
            if col_idx in [4, 5, 6, 10, 11] and isinstance(val, (int, float)):
                cell.number_format = '#,##0'
        
        # 评级颜色
        rating = scenario.get('rating', '')
        rating_cell = ws.cell(row=idx, column=13)
        if "★★★★★" in rating:
            rating_cell.fill = GREEN_FILL
            rating_cell.font = GREEN_FONT
        elif "★★★★" in rating:
            rating_cell.fill = YELLOW_FILL
            rating_cell.font = YELLOW_FONT
        elif "★★★" in rating:
            rating_cell.fill = PatternFill("solid", fgColor="FCE4D6")
            rating_cell.font = Font(name="Microsoft YaHei", size=9, bold=True, color="BF4000")
    
    # 设置列宽
    ws.column_dimensions["A"].width = 18
    for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]:
        ws.column_dimensions[col].width = 12
    
    wb.save(output_path)
    print(f"己方视角ROI分析已保存：{output_path}")


if __name__ == '__main__':
    # 示例用法
    customer_scenarios = [
        {"name": "智慧工地", "customer_investment": 200, "customer_annual_benefit": 500, "payback_period": "4-8个月", "roi_5year": "300%", "rating": "★★★★★"},
        {"name": "两票三制管理", "customer_investment": 80, "customer_annual_benefit": 65, "payback_period": "1.2-1.5年", "roi_5year": "300%", "rating": "★★★★★"},
        {"name": "虚拟电厂", "customer_investment": 200, "customer_annual_benefit": 350, "payback_period": "1.5-2年", "roi_5year": "—", "rating": "★★★★☆"},
    ]
    
    vendor_scenarios = [
        {"name": "智慧工地", "rd_team_size": 8, "rd_duration": 4, "rd_investment": 80, 
         "delivery_cost": 30, "quote": 200, "gross_margin": "85%", "reusable_ratio": "70%",
         "target_customers_3y": 30, "revenue_3y": 6000, "profit_3y": 4620, "payback_period": "0.5个月", "rating": "★★★★★"},
        {"name": "两票三制管理", "rd_team_size": 5, "rd_duration": 3, "rd_investment": 37.5, 
         "delivery_cost": 10, "quote": 80, "gross_margin": "88%", "reusable_ratio": "65%",
         "target_customers_3y": 50, "revenue_3y": 4000, "profit_3y": 3290, "payback_period": "0.4个月", "rating": "★★★★★"},
    ]
    
    create_customer_roi_excel(customer_scenarios, "客户视角ROI示例.xlsx")
    create_vendor_roi_excel(vendor_scenarios, "己方视角ROI示例.xlsx")
