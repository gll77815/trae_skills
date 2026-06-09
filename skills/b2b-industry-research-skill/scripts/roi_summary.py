#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成客户视角ROI汇总Excel

功能：
1. 汇总各场景的客户视角ROI数据
2. 计算投资回收期、5年ROI等关键指标
3. 按ROI分级（★★★★★ / ★★★★☆ / ★★★☆☆ / ★★☆☆☆）
4. 生成4梯队推荐（强推/推荐/可选/观望）

使用时机：Phase 5.1 - 客户视角ROI分析
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from datetime import datetime
import re


# ROI分级标准
RATING_CRITERIA = {
    "★★★★★": {"payback_months": (0, 12), "roi_5year": (300, float('inf'))},
    "★★★★☆": {"payback_months": (12, 18), "roi_5year": (200, 300)},
    "★★★☆☆": {"payback_months": (18, 24), "roi_5year": (100, 200)},
    "★★☆☆☆": {"payback_months": (24, 36), "roi_5year": (50, 100)},
    "★☆☆☆☆": {"payback_months": (36, float('inf')), "roi_5year": (0, 50)},
}


def parse_payback_period(payback_str):
    """
    解析回收期字符串，返回月数
    
    支持格式：
    - "12个月" -> 12
    - "1.5年" -> 18
    - "1年3个月" -> 15
    """
    if not payback_str:
        return None
    
    payback_str = str(payback_str).strip()
    
    # 尝试提取数字
    numbers = re.findall(r'[0-9.]+', payback_str)
    if not numbers:
        return None
    
    value = float(numbers[0])
    
    # 根据单位转换
    if '年' in payback_str and '个月' in payback_str:
        # 处理 "1年3个月" 格式
        years = value
        months_match = re.search(r'(\d+)个月', payback_str)
        months = int(months_match.group(1)) if months_match else 0
        return years * 12 + months
    elif '年' in payback_str:
        return value * 12
    else:
        return value


def parse_roi_value(roi_str):
    """
    解析ROI字符串，返回百分比数值
    
    支持格式：
    - "300%" -> 300
    - "3倍" -> 300
    - "3x" -> 300
    """
    if not roi_str:
        return None
    
    roi_str = str(roi_str).strip()
    
    # 提取数字
    numbers = re.findall(r'[0-9.]+', roi_str)
    if not numbers:
        return None
    
    value = float(numbers[0])
    
    # 根据单位转换
    if '倍' in roi_str or 'x' in roi_str.lower():
        return value * 100
    else:
        return value


def calculate_rating(payback_months, roi_5year):
    """
    根据回收期和ROI计算评级
    
    返回:
        str: 星级评级
    """
    # 优先使用回收期判断
    if payback_months is not None:
        for rating, criteria in RATING_CRITERIA.items():
            min_months, max_months = criteria["payback_months"]
            if min_months <= payback_months < max_months:
                return rating
    
    # 备用使用ROI判断
    if roi_5year is not None:
        for rating, criteria in RATING_CRITERIA.items():
            min_roi, max_roi = criteria["roi_5year"]
            if min_roi <= roi_5year < max_roi:
                return rating
    
    return "待评估"


def get_recommendation_level(rating):
    """
    根据评级返回推荐梯队
    """
    mapping = {
        "★★★★★": "强推",
        "★★★★☆": "推荐",
        "★★★☆☆": "可选",
        "★★☆☆☆": "观望",
        "★☆☆☆☆": "观望",
        "待评估": "待评估",
    }
    return mapping.get(rating, "待评估")


def create_roi_summary_excel(scenes_roi_data, industry, output_path=None):
    """
    创建客户视角ROI汇总Excel
    
    参数:
        scenes_roi_data: 场景ROI数据列表
        industry: 行业名称
        output_path: 输出路径（可选）
    
    返回:
        生成的文件路径
    """
    
    wb = openpyxl.Workbook()
    
    # ===== Sheet 1: ROI主表 =====
    ws_main = wb.active
    ws_main.title = "客户视角ROI汇总"
    
    # 定义样式
    fill_header = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    fill_strong = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 绿色-强推
    fill_recommend = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # 浅绿-推荐
    fill_optional = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # 黄色-可选
    fill_watch = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 红色-观望
    
    font_header = Font(bold=True, color="FFFFFF", size=11)
    font_bold = Font(bold=True, size=10)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # 写入标题
    headers = ["场景名称", "客户投入", "年收益", "回收期(月)", "5年ROI(%)", "评级", "推荐梯队"]
    
    for col, header in enumerate(headers, 1):
        cell = ws_main.cell(row=1, column=col, value=header)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_center
        cell.border = thin_border
    
    # 处理数据并写入
    processed_data = []
    for scene in scenes_roi_data:
        payback_months = parse_payback_period(scene.get("payback_period", ""))
        roi_5year = parse_roi_value(scene.get("five_year_roi", ""))
        rating = calculate_rating(payback_months, roi_5year)
        recommendation = get_recommendation_level(rating)
        
        processed_data.append({
            "scene_name": scene.get("scene_name", ""),
            "customer_investment": scene.get("customer_investment", ""),
            "annual_benefit": scene.get("annual_benefit", ""),
            "payback_months": payback_months,
            "roi_5year": roi_5year,
            "rating": rating,
            "recommendation": recommendation,
        })
    
    # 按评级排序（强推在前）
    rating_order = {"★★★★★": 1, "★★★★☆": 2, "★★★☆☆": 3, "★★☆☆☆": 4, "★☆☆☆☆": 5, "待评估": 6}
    processed_data.sort(key=lambda x: rating_order.get(x["rating"], 99))
    
    # 写入数据
    for row_idx, data in enumerate(processed_data, 2):
        ws_main.cell(row=row_idx, column=1, value=data["scene_name"]).alignment = align_left
        ws_main.cell(row=row_idx, column=2, value=data["customer_investment"]).alignment = align_left
        ws_main.cell(row=row_idx, column=3, value=data["annual_benefit"]).alignment = align_left
        
        cell_payback = ws_main.cell(row=row_idx, column=4, value=data["payback_months"])
        cell_payback.alignment = align_center
        
        cell_roi = ws_main.cell(row=row_idx, column=5, value=data["roi_5year"])
        cell_roi.alignment = align_center
        
        ws_main.cell(row=row_idx, column=6, value=data["rating"]).alignment = align_center
        
        cell_rec = ws_main.cell(row=row_idx, column=7, value=data["recommendation"])
        cell_rec.alignment = align_center
        
        # 根据推荐梯队设置背景色
        if data["recommendation"] == "强推":
            cell_rec.fill = fill_strong
        elif data["recommendation"] == "推荐":
            cell_rec.fill = fill_recommend
        elif data["recommendation"] == "可选":
            cell_rec.fill = fill_optional
        elif data["recommendation"] == "观望":
            cell_rec.fill = fill_watch
        
        # 应用边框
        for col in range(1, 8):
            ws_main.cell(row=row_idx, column=col).border = thin_border
    
    # 设置列宽
    ws_main.column_dimensions["A"].width = 30
    ws_main.column_dimensions["B"].width = 20
    ws_main.column_dimensions["C"].width = 20
    ws_main.column_dimensions["D"].width = 12
    ws_main.column_dimensions["E"].width = 12
    ws_main.column_dimensions["F"].width = 12
    ws_main.column_dimensions["G"].width = 12
    
    # 冻结首行
    ws_main.freeze_panes = "A2"
    
    # ===== Sheet 2: 梯队分析 =====
    ws_analysis = wb.create_sheet("梯队分析")
    
    # 统计各梯队场景数
    from collections import Counter
    recommendation_counts = Counter([d["recommendation"] for d in processed_data])
    
    analysis_data = [
        ["推荐梯队", "场景数量", "占比", "策略建议"],
        ["强推", recommendation_counts.get("强推", 0), 
         f"{recommendation_counts.get('强推', 0) / len(processed_data) * 100:.1f}%" if processed_data else "0%",
         "优先投入，快速占领市场"],
        ["推荐", recommendation_counts.get("推荐", 0),
         f"{recommendation_counts.get('推荐', 0) / len(processed_data) * 100:.1f}%" if processed_data else "0%",
         "重点跟进，争取转化为强推"],
        ["可选", recommendation_counts.get("可选", 0),
         f"{recommendation_counts.get('可选', 0) / len(processed_data) * 100:.1f}%" if processed_data else "0%",
         "视资源情况选择性投入"],
        ["观望", recommendation_counts.get("观望", 0),
         f"{recommendation_counts.get('观望', 0) / len(processed_data) * 100:.1f}%" if processed_data else "0%",
         "暂不投入，持续跟踪市场变化"],
    ]
    
    for row_idx, row_data in enumerate(analysis_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_analysis.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.fill = fill_header
                cell.font = font_header
            cell.alignment = align_center
            cell.border = thin_border
    
    ws_analysis.column_dimensions["A"].width = 12
    ws_analysis.column_dimensions["B"].width = 12
    ws_analysis.column_dimensions["C"].width = 12
    ws_analysis.column_dimensions["D"].width = 30
    
    # ===== Sheet 3: 使用说明 =====
    ws_help = wb.create_sheet("使用说明")
    help_content = [
        [f"{industry}行业 - 客户视角ROI分析", ""],
        ["", ""],
        ["评级标准", ""],
        ["★★★★★ 强推", "回收期 ≤ 12个月 或 5年ROI ≥ 300%"],
        ["★★★★☆ 推荐", "回收期 12-18个月 或 5年ROI 200-300%"],
        ["★★★☆☆ 可选", "回收期 18-24个月 或 5年ROI 100-200%"],
        ["★★☆☆☆ 观望", "回收期 24-36个月 或 5年ROI 50-100%"],
        ["★☆☆☆☆ 观望", "回收期 > 36个月 或 5年ROI < 50%"],
        ["", ""],
        ["数据说明", ""],
        ["客户投入", "客户采购方案所需的一次性投入"],
        ["年收益", "客户使用方案后每年获得的收益/节省"],
        ["回收期", "客户收回投资所需的时间（月）"],
        ["5年ROI", "客户5年内获得的投资回报率"],
        ["", ""],
        ["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
    ]
    
    for row_idx, row_data in enumerate(help_content, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_help.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
            elif row_idx == 3 or row_idx == 10:
                cell.font = Font(bold=True, size=11)
    
    ws_help.column_dimensions["A"].width = 20
    ws_help.column_dimensions["B"].width = 50
    
    # 保存文件
    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d")
        output_path = f"{industry}行业_场景投入产出比汇总对比_{timestamp}.xlsx"
    
    wb.save(output_path)
    print(f"ROI汇总Excel已生成: {output_path}")
    return output_path


def main():
    """示例用法"""
    # 示例数据
    scenes_roi = [
        {
            "scene_name": "智能巡检",
            "customer_investment": "50万元",
            "annual_benefit": "节省人工成本80万元/年",
            "payback_period": "8个月",
            "five_year_roi": "700%",
        },
        {
            "scene_name": "设备预测性维护",
            "customer_investment": "120万元",
            "annual_benefit": "减少停机损失200万元/年",
            "payback_period": "7个月",
            "five_year_roi": "733%",
        },
        {
            "scene_name": "能源管理系统",
            "customer_investment": "80万元",
            "annual_benefit": "节省能源成本30万元/年",
            "payback_period": "32个月",
            "five_year_roi": "87%",
        },
    ]
    
    create_roi_summary_excel(scenes_roi, "示例行业", "示例_ROI汇总.xlsx")


if __name__ == "__main__":
    main()
